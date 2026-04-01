from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "副業_経費管理テンプレート.xlsx"


def style_range(sheet, start_row: int, end_row: int, start_col: int, end_col: int, border: Border) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).border = border


def next_available_output(base_path: Path) -> Path:
    if not base_path.exists():
        return base_path

    stem = base_path.stem
    suffix = base_path.suffix
    counter = 2
    while True:
        candidate = base_path.with_name(f"{stem}_{counter}{suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "経費入力"
    summary = wb.create_sheet("月別集計")
    planner = wb.create_sheet("経費候補整理")
    master = wb.create_sheet("マスタ")
    guide = wb.create_sheet("使い方")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    master_headers = {
        "A1": "経費区分",
        "B1": "支払方法",
        "C1": "証憑状況",
        "D1": "補足メモ",
    }
    for cell, value in master_headers.items():
        master[cell] = value
        master[cell].fill = header_fill
        master[cell].font = header_font
        master[cell].border = border

    categories = [
        "通信費",
        "消耗品費",
        "広告宣伝費",
        "旅費交通費",
        "車両費",
        "接待交際費",
        "新聞図書費",
        "外注費",
        "支払手数料",
        "地代家賃",
        "水道光熱費",
        "保険料",
        "修繕費",
        "減価償却費",
        "研修費",
        "会議費",
        "ソフトウェア利用料",
        "荷造運賃",
        "租税公課",
        "雑費",
    ]
    payment_methods = ["現金", "クレジットカード", "銀行振込", "口座引落", "電子マネー", "その他"]
    receipt_statuses = ["あり", "電子保存あり", "再発行待ち", "なし"]
    notes = [
        "個人事業主の一般的な経費項目例です",
        "必要に応じてマスタを編集して項目追加できます",
        "証憑の保管場所をそろえると申告前の確認が楽です",
    ]

    for i, item in enumerate(categories, start=2):
        master[f"A{i}"] = item
    for i, item in enumerate(payment_methods, start=2):
        master[f"B{i}"] = item
    for i, item in enumerate(receipt_statuses, start=2):
        master[f"C{i}"] = item
    for i, item in enumerate(notes, start=2):
        master[f"D{i}"] = item

    style_range(master, 1, max(len(categories), len(payment_methods), len(receipt_statuses), len(notes)) + 1, 1, 4, border)

    wb.defined_names.add(DefinedName("ExpenseCategories", attr_text=f"'マスタ'!$A$2:$A${len(categories) + 1}"))
    wb.defined_names.add(DefinedName("PaymentMethods", attr_text="'マスタ'!$B$2:$B$7"))
    wb.defined_names.add(DefinedName("ReceiptStatuses", attr_text="'マスタ'!$C$2:$C$5"))

    guide["A1"] = "副業の経費管理テンプレート 使い方"
    guide["A1"].fill = header_fill
    guide["A1"].font = header_font
    guide["A5"] = "入力のコツ"
    guide["A5"].fill = header_fill
    guide["A5"].font = header_font
    guide["A2"] = "1. 「経費入力」シートに、支払日ごとに1行ずつ入力します。"
    guide["A3"] = "2. 家事按分があるものは「事業利用割合(%)」を入力すると、按分後金額が自動計算されます。"
    guide["A4"] = "3. 領収書や請求書の保管場所を「証憑保管メモ」に記録しておくと、確定申告時の確認が楽です。"
    guide["A6"] = "・勘定科目に迷ったものは近い区分で入力し、申告前に最終確認してください。"
    guide["A7"] = "・私用と共用の支出は、根拠が説明できる割合で按分してください。"
    guide["A8"] = "・「申告チェック」列で未整理の行を見つけられます。"
    guide["A9"] = "・「経費候補整理」シートに、住居費・自動車・バイク・PC・スマホ・ネットの候補を整理しています。"
    guide["A10"] = "このテンプレートは整理用です。最終的な税務判断は最新制度や専門家確認を前提にしてください。"
    guide.column_dimensions["A"].width = 92
    for row in range(1, 11):
        guide[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "No",
        "支払日",
        "月",
        "取引先/支払先",
        "内容",
        "経費区分",
        "支払方法",
        "税込金額",
        "事業利用割合(%)",
        "事業按分後金額",
        "消費税メモ",
        "証憑状況",
        "証憑保管メモ",
        "支払済",
        "申告チェック",
        "備考",
    ]
    widths = [8, 12, 10, 22, 28, 18, 14, 14, 16, 16, 14, 12, 26, 10, 16, 24]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    for row in range(2, 302):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=3, value=f'=TEXT(B{row},"yyyy-mm")')
        ws.cell(row=row, column=9, value=100)
        ws.cell(row=row, column=10, value=f'=IF(OR(H{row}="",I{row}=""),"",ROUND(H{row}*I{row}/100,0))')
        ws.cell(row=row, column=14, value="済")
        ws.cell(
            row=row,
            column=15,
            value=f'=IF(OR(B{row}="",F{row}="",J{row}="",L{row}=""),"要確認",IF(L{row}="なし","証憑不足","OK"))',
        )
        for col in range(1, 17):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = left

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:P301"

    category_dv = DataValidation(type="list", formula1="=ExpenseCategories", allow_blank=True)
    payment_dv = DataValidation(type="list", formula1="=PaymentMethods", allow_blank=True)
    receipt_dv = DataValidation(type="list", formula1="=ReceiptStatuses", allow_blank=True)
    paid_dv = DataValidation(type="list", formula1='"済,未"', allow_blank=True)
    percent_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)

    for dv in [category_dv, payment_dv, receipt_dv, paid_dv, percent_dv]:
        ws.add_data_validation(dv)

    category_dv.add("F2:F301")
    payment_dv.add("G2:G301")
    receipt_dv.add("L2:L301")
    paid_dv.add("N2:N301")
    percent_dv.add("I2:I301")

    for row in range(2, 302):
        ws[f"B{row}"].number_format = "yyyy-mm-dd"
        ws[f"H{row}"].number_format = "#,##0"
        ws[f"I{row}"].number_format = "0"
        ws[f"J{row}"].number_format = "#,##0"

    warning_rule = FormulaRule(formula=['$O2="要確認"'], fill=warn_fill)
    ws.conditional_formatting.add("A2:P301", warning_rule)

    summary["A1"] = "副業経費 月別サマリー"
    summary["A1"].fill = header_fill
    summary["A1"].font = header_font
    summary["A3"] = "月"
    summary["B3"] = "事業按分後合計"
    summary["D3"] = "経費区分"
    summary["E3"] = "事業按分後合計"

    for cell in ["A3", "B3", "D3", "E3"]:
        summary[cell].fill = sub_fill
        summary[cell].font = bold_font
        summary[cell].border = border

    months = [f"2026-{month:02d}" for month in range(1, 13)]
    for idx, month in enumerate(months, start=4):
        summary[f"A{idx}"] = month
        summary[f"B{idx}"] = f'=SUMIF(経費入力!$C$2:$C$301,A{idx},経費入力!$J$2:$J$301)'
        summary[f"A{idx}"].border = border
        summary[f"B{idx}"].border = border
        summary[f"B{idx}"].number_format = "#,##0"

    for idx, category in enumerate(categories, start=4):
        summary[f"D{idx}"] = category
        summary[f"E{idx}"] = f'=SUMIF(経費入力!$F$2:$F$301,D{idx},経費入力!$J$2:$J$301)'
        summary[f"D{idx}"].border = border
        summary[f"E{idx}"].border = border
        summary[f"E{idx}"].number_format = "#,##0"

    summary["A18"] = "年間合計"
    summary["B18"] = "=SUM(B4:B15)"
    summary["D21"] = "未確認件数"
    summary["E21"] = '=COUNTIF(経費入力!$O$2:$O$301,"要確認")'
    for cell in ["A18", "B18", "D21", "E21"]:
        summary[cell].fill = accent_fill
        summary[cell].font = bold_font
        summary[cell].border = border

    chart = PieChart()
    labels = Reference(summary, min_col=4, min_row=4, max_row=3 + len(categories))
    data = Reference(summary, min_col=5, min_row=3, max_row=3 + len(categories))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "経費区分ごとの構成比"
    chart.height = 9
    chart.width = 13
    summary.add_chart(chart, "G3")

    summary.column_dimensions["A"].width = 12
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["D"].width = 18
    summary.column_dimensions["E"].width = 16

    planner["A1"] = "副業の主要経費候補 整理シート"
    planner["A1"].fill = header_fill
    planner["A1"].font = header_font
    planner["A2"] = "住居費・車両費・通信費など、事業利用分だけを按分して整理するための表です。"
    planner["A3"] = "按分割合は、面積・使用時間・走行記録など、説明できる根拠に沿って入力してください。"
    planner["A4"] = "住宅ローン元本などは通常の必要経費整理になじみにくいため、この表では注意喚起を入れています。"

    planner_headers = [
        "大分類",
        "費用項目",
        "一般的な勘定科目",
        "支払頻度",
        "年間支払見込(円)",
        "事業利用割合(%)",
        "年間経費見込(円)",
        "整理のポイント",
        "証憑・記録",
        "入力メモ",
    ]
    for col, header in enumerate(planner_headers, start=1):
        cell = planner.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    planner_rows = [
        (
            "自宅をオフィス利用",
            "家賃",
            "地代家賃",
            "毎月",
            "",
            "",
            '=IF(OR(E7="",F7=""),"",ROUND(E7*F7/100,0))',
            "仕事専用スペースの面積や使用時間で按分。事業利用部分のみ。",
            "賃貸契約書、家賃明細、按分根拠メモ",
            "",
        ),
        (
            "自宅をオフィス利用",
            "共益費・管理費",
            "地代家賃",
            "毎月",
            "",
            "",
            '=IF(OR(E8="",F8=""),"",ROUND(E8*F8/100,0))',
            "家賃と同じ按分根拠で整理しやすい項目です。",
            "請求明細、按分根拠メモ",
            "",
        ),
        (
            "自宅をオフィス利用",
            "電気代",
            "水道光熱費",
            "毎月",
            "",
            "",
            '=IF(OR(E9="",F9=""),"",ROUND(E9*F9/100,0))',
            "仕事で使う部屋や使用時間に応じて按分。",
            "請求書、利用明細",
            "",
        ),
        (
            "自宅をオフィス利用",
            "ガス代",
            "水道光熱費",
            "毎月",
            "",
            "",
            '=IF(OR(E10="",F10=""),"",ROUND(E10*F10/100,0))',
            "業務との関係が弱い場合は対象外になりやすいので慎重に判断。",
            "請求書、按分根拠メモ",
            "",
        ),
        (
            "自宅をオフィス利用",
            "水道代",
            "水道光熱費",
            "毎月",
            "",
            "",
            '=IF(OR(E11="",F11=""),"",ROUND(E11*F11/100,0))',
            "業務との関係が説明できる範囲で按分。",
            "請求書、按分根拠メモ",
            "",
        ),
        (
            "自宅をオフィス利用",
            "火災保険・地震保険の事業利用分",
            "保険料",
            "年1回",
            "",
            "",
            '=IF(OR(E12="",F12=""),"",ROUND(E12*F12/100,0))',
            "自宅兼事務所で使う部分のみ。契約内容の確認推奨。",
            "保険証券、按分根拠",
            "",
        ),
        (
            "自宅をオフィス利用",
            "住宅ローン利息など要個別判断項目",
            "要確認",
            "毎月",
            "",
            "",
            '=IF(OR(E13="",F13=""),"",ROUND(E13*F13/100,0))',
            "住宅ローン元本は通常経費整理になじみません。利息等は個別判断前提。",
            "返済予定表、相談メモ",
            "",
        ),
        (
            "自動車",
            "ガソリン代",
            "車両費",
            "毎月",
            "",
            "",
            '=IF(OR(E14="",F14=""),"",ROUND(E14*F14/100,0))',
            "業務走行距離の割合で按分すると整理しやすいです。",
            "レシート、走行記録",
            "",
        ),
        (
            "自動車",
            "駐車場代",
            "車両費",
            "毎月",
            "",
            "",
            '=IF(OR(E15="",F15=""),"",ROUND(E15*F15/100,0))',
            "仕事用専用駐車場なら按分不要のこともあります。",
            "契約書、領収書",
            "",
        ),
        (
            "自動車",
            "高速代・有料道路",
            "旅費交通費",
            "都度",
            "",
            "",
            '=IF(OR(E16="",F16=""),"",ROUND(E16*F16/100,0))',
            "業務利用分を個別記録で区分しやすい項目です。",
            "利用履歴、ETC明細",
            "",
        ),
        (
            "自動車",
            "車検・修理・メンテナンス",
            "修繕費",
            "随時",
            "",
            "",
            '=IF(OR(E17="",F17=""),"",ROUND(E17*F17/100,0))',
            "私用兼用なら業務割合で按分。",
            "整備明細、請求書",
            "",
        ),
        (
            "自動車",
            "自動車保険",
            "保険料",
            "年1回",
            "",
            "",
            '=IF(OR(E18="",F18=""),"",ROUND(E18*F18/100,0))',
            "業務使用割合に応じて按分。",
            "保険証券、支払明細",
            "",
        ),
        (
            "自動車",
            "自動車税・重量税など",
            "租税公課",
            "年1回",
            "",
            "",
            '=IF(OR(E19="",F19=""),"",ROUND(E19*F19/100,0))',
            "業務使用分の整理前提。車両取得時の扱いは別途確認。",
            "納税通知書、支払記録",
            "",
        ),
        (
            "自動車",
            "車両購入費",
            "減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E20="",F20=""),"",ROUND(E20*F20/100,0))',
            "購入額全額を一度に経費化せず、減価償却になる場合があります。",
            "売買契約書、領収書",
            "",
        ),
        (
            "バイク",
            "ガソリン代",
            "車両費",
            "毎月",
            "",
            "",
            '=IF(OR(E21="",F21=""),"",ROUND(E21*F21/100,0))',
            "配達・訪問など業務走行分の割合で按分。",
            "レシート、走行記録",
            "",
        ),
        (
            "バイク",
            "駐輪場代",
            "車両費",
            "毎月",
            "",
            "",
            '=IF(OR(E22="",F22=""),"",ROUND(E22*F22/100,0))',
            "業務用なら区分しやすい項目です。",
            "契約書、領収書",
            "",
        ),
        (
            "バイク",
            "修理・メンテナンス",
            "修繕費",
            "随時",
            "",
            "",
            '=IF(OR(E23="",F23=""),"",ROUND(E23*F23/100,0))',
            "私用兼用なら業務割合で按分。",
            "整備明細、請求書",
            "",
        ),
        (
            "バイク",
            "保険・税金",
            "保険料 / 租税公課",
            "年1回",
            "",
            "",
            '=IF(OR(E24="",F24=""),"",ROUND(E24*F24/100,0))',
            "保険と税金は分けて経費入力すると集計しやすいです。",
            "保険証券、納税記録",
            "",
        ),
        (
            "バイク",
            "車体購入費",
            "減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E25="",F25=""),"",ROUND(E25*F25/100,0))',
            "購入額により減価償却を検討。",
            "売買契約書、領収書",
            "",
        ),
        (
            "パソコン",
            "本体購入費",
            "消耗品費 / 減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E26="",F26=""),"",ROUND(E26*F26/100,0))',
            "取得価額が小さい場合と大きい場合で処理が分かれやすい項目です。",
            "領収書、購入明細",
            "",
        ),
        (
            "パソコン",
            "モニター・キーボード等周辺機器",
            "消耗品費 / 減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E27="",F27=""),"",ROUND(E27*F27/100,0))',
            "1単位ごとの取得価額で判定されることがあります。",
            "領収書、購入明細",
            "",
        ),
        (
            "パソコン",
            "修理費",
            "修繕費",
            "随時",
            "",
            "",
            '=IF(OR(E28="",F28=""),"",ROUND(E28*F28/100,0))',
            "通常の維持管理なら修繕費として整理しやすいです。",
            "修理明細、請求書",
            "",
        ),
        (
            "パソコン",
            "ソフトウェア・クラウド利用料",
            "ソフトウェア利用料",
            "毎月",
            "",
            "",
            '=IF(OR(E29="",F29=""),"",ROUND(E29*F29/100,0))',
            "仕事専用なら按分不要のことがあります。",
            "請求書、契約情報",
            "",
        ),
        (
            "スマホ",
            "端末購入費",
            "消耗品費 / 減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E30="",F30=""),"",ROUND(E30*F30/100,0))',
            "事業利用割合を通話履歴や使用状況で説明できるように。",
            "領収書、購入明細",
            "",
        ),
        (
            "スマホ",
            "基本料金・通話料",
            "通信費",
            "毎月",
            "",
            "",
            '=IF(OR(E31="",F31=""),"",ROUND(E31*F31/100,0))',
            "私用兼用なら按分前提で整理。",
            "利用明細、通話履歴",
            "",
        ),
        (
            "スマホ",
            "業務アプリ課金",
            "通信費 / ソフトウェア利用料",
            "毎月",
            "",
            "",
            '=IF(OR(E32="",F32=""),"",ROUND(E32*F32/100,0))',
            "仕事目的が明確なら整理しやすい項目です。",
            "請求書、課金履歴",
            "",
        ),
        (
            "ネット",
            "自宅インターネット回線",
            "通信費",
            "毎月",
            "",
            "",
            '=IF(OR(E33="",F33=""),"",ROUND(E33*F33/100,0))',
            "業務利用時間や利用者数などで按分の根拠を残すと安心です。",
            "請求明細、按分根拠メモ",
            "",
        ),
        (
            "ネット",
            "Wi-Fiルーター・中継器購入",
            "消耗品費 / 減価償却費",
            "購入時",
            "",
            "",
            '=IF(OR(E34="",F34=""),"",ROUND(E34*F34/100,0))',
            "取得価額に応じて処理方法が変わる可能性があります。",
            "領収書、購入明細",
            "",
        ),
        (
            "ネット",
            "工事費・初期費用",
            "通信費 / 繰延的に要確認",
            "契約時",
            "",
            "",
            '=IF(OR(E35="",F35=""),"",ROUND(E35*F35/100,0))',
            "契約内容により処理が分かれることがあるためメモ推奨。",
            "契約書、請求書",
            "",
        ),
    ]

    for idx, row_data in enumerate(planner_rows, start=7):
        for col, value in enumerate(row_data, start=1):
            planner.cell(row=idx, column=col, value=value)
            planner.cell(row=idx, column=col).border = border
            planner.cell(row=idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        planner[f"E{idx}"].number_format = "#,##0"
        planner[f"F{idx}"].number_format = "0"
        planner[f"G{idx}"].number_format = "#,##0"

    planner["E36"] = "年間経費見込 合計"
    planner["E36"].fill = accent_fill
    planner["E36"].font = bold_font
    planner["F36"] = ""
    planner["G36"] = "=SUM(G7:G35)"
    planner["G36"].fill = accent_fill
    planner["G36"].font = bold_font
    planner["G36"].number_format = "#,##0"
    planner["I36"] = "使い方"
    planner["J36"] = "このシートで見積もったら、実支払分を「経費入力」へ転記"
    planner["I36"].fill = accent_fill
    planner["I36"].font = bold_font
    planner["J36"].fill = accent_fill

    planner.freeze_panes = "A6"
    planner.auto_filter.ref = "A6:J35"
    planner.column_dimensions["A"].width = 18
    planner.column_dimensions["B"].width = 26
    planner.column_dimensions["C"].width = 22
    planner.column_dimensions["D"].width = 10
    planner.column_dimensions["E"].width = 16
    planner.column_dimensions["F"].width = 14
    planner.column_dimensions["G"].width = 16
    planner.column_dimensions["H"].width = 38
    planner.column_dimensions["I"].width = 28
    planner.column_dimensions["J"].width = 24

    planner["F7"] = 20
    planner["F8"] = 20
    planner["F9"] = 20
    planner["F10"] = 10
    planner["F11"] = 10
    planner["F12"] = 20
    planner["F14"] = 30
    planner["F15"] = 100
    planner["F16"] = 100
    planner["F17"] = 30
    planner["F18"] = 30
    planner["F19"] = 30
    planner["F20"] = 30
    planner["F21"] = 30
    planner["F22"] = 100
    planner["F23"] = 30
    planner["F24"] = 30
    planner["F25"] = 30
    planner["F26"] = 80
    planner["F27"] = 80
    planner["F28"] = 80
    planner["F29"] = 100
    planner["F30"] = 50
    planner["F31"] = 50
    planner["F32"] = 100
    planner["F33"] = 50
    planner["F34"] = 50
    planner["F35"] = 50

    estimated_amounts = {
        "E7": 1200000,
        "E8": 180000,
        "E9": 144000,
        "E10": 60000,
        "E11": 48000,
        "E12": 20000,
        "E13": 240000,
        "E14": 120000,
        "E15": 180000,
        "E16": 36000,
        "E17": 80000,
        "E18": 70000,
        "E19": 45000,
        "E20": 3000000,
        "E21": 36000,
        "E22": 36000,
        "E23": 30000,
        "E24": 20000,
        "E25": 500000,
        "E26": 180000,
        "E27": 60000,
        "E28": 20000,
        "E29": 60000,
        "E30": 120000,
        "E31": 96000,
        "E32": 12000,
        "E33": 72000,
        "E34": 15000,
        "E35": 20000,
    }
    for cell, value in estimated_amounts.items():
        planner[cell] = value
        planner[cell].number_format = "#,##0"

    sample = {
        "B2": date(2026, 4, 1),
        "D2": "X社",
        "E2": "開業準備の書籍購入",
        "F2": "新聞図書費",
        "G2": "クレジットカード",
        "H2": 2800,
        "I2": 100,
        "K2": "必要ならインボイス該当有無もメモ",
        "L2": "電子保存あり",
        "M2": "Google Drive > 経費 > 2026-04",
        "N2": "済",
        "P2": "開業前支出は開業費扱いも要検討",
    }
    for cell, value in sample.items():
        ws[cell] = value
    ws["C2"] = '=TEXT(B2,"yyyy-mm")'
    ws["J2"] = '=IF(OR(H2="",I2=""),"",ROUND(H2*I2/100,0))'
    ws["O2"] = '=IF(OR(B2="",F2="",J2="",L2=""),"要確認",IF(L2="なし","証憑不足","OK"))'
    ws["B2"].number_format = "yyyy-mm-dd"
    ws["H2"].number_format = "#,##0"
    ws["I2"].number_format = "0"
    ws["J2"].number_format = "#,##0"

    for sheet in [ws, summary, planner, master, guide]:
        sheet.sheet_view.showGridLines = True

    master.sheet_state = "hidden"

    output_path = OUTPUT
    try:
        wb.save(output_path)
    except PermissionError:
        output_path = ROOT / "副業_経費管理テンプレート_更新版.xlsx"
        try:
            wb.save(output_path)
        except PermissionError:
            output_path = next_available_output(output_path)
            wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
